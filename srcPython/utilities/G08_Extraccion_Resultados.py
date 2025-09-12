import comtypes.client

import pandas as pd
import numpy as np

import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class Extraccion_Resultados:
    def __init__(self):
        self.SapModel, self.ETABSObject, self.helper = self.Connect_ETABS()
        self.Excel_tables()
    
    def Connect_ETABS(self):
        helper = comtypes.client.CreateObject('ETABSv1.Helper')                 # Maneja las conexiones con ETABS
        helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)            # Accede a métodos avanzados para interactuar con ETABS
        ETABSObject = helper.GetObject("CSI.ETABS.API.ETABSObject")             # Se conecta a ETABS 
        SapModel = ETABSObject.SapModel                                         # Accede al modelo estructural actualmente cargado en ETABS
        return SapModel, ETABSObject, helper
    
    def Modos_vibracion(self):
        # Filtrado de tabla
        ret = self.SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = self.SapModel.Results.Setup.SetCaseSelectedForOutput("Modal")
        
        # Extracción de las masas participativas de los modos de vibración
        NumberResults = 0; Period = []; Ux = []; Uy = []; R_z = []
        [NumberResults, _, _, _, Period, Ux, Uy, _, _, _, _, _, _, R_z, _, _, _, ret] = self.SapModel.Results.ModalParticipatingMassRatios(NumberResults, [], [], [], Period, Ux, Uy, [], [], [], [], [], [], R_z, [], [], [])
        
        # Creacion de tabla resumen
        Mode = np.arange(1, NumberResults + 1, 1)
        df = pd.DataFrame({'Modo': Mode, 'Periodo': Period, 'Ux': Ux, 'Uy': Uy, 'Rz': R_z})
        df = df.iloc[:3]
        max_columns = df[['Ux', 'Uy', 'Rz']].idxmax(axis=1)
        
        for index, column in max_columns.items():
            if column == 'Ux':      type_mode = 'MODO TRANSLACIONAL X'
            elif column == 'Uy':    type_mode = 'MODO TRANSLACIONAL Y'
            elif column == 'Rz':    type_mode = 'MODO ROTACIONAL'
            number_mode = str(df.at[index, 'Modo'])
            print(number_mode +'° modo vibracional : T'+ number_mode +' = '+ str(df.at[index, 'Periodo'].round(3)) +'s  ----> '+ type_mode +'\n')
    
    def k_value(self, Tp, Tl):
        # Filtrado de tabla
        ret = self.SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = self.SapModel.Results.Setup.SetCaseSelectedForOutput("Modal")
        
        # Extracción de los periodos
        NumberResults = 0; Period = []; Ux = []; Uy = []; R_z = []
        [NumberResults, _, _, _, Period, Ux, Uy, _, _, _, _, _, _, R_z, _, _, _, ret] = self.SapModel.Results.ModalParticipatingMassRatios(NumberResults, [], [], [], Period, Ux, Uy, [], [], [], [], [], [], R_z, [], [], [])
        
        # Creacion de tabla resumen
        df = pd.DataFrame({'Periodo': Period, 'Ux': Ux, 'Uy': Uy, 'Rz': R_z})
        df = df.iloc[:3]
        max_columns = df[['Ux', 'Uy', 'Rz']].idxmax(axis=1)

        # Calculo de valores sismicos
        for index, column in max_columns.items():
            if column != 'Rz':
                if column == 'Ux':      T = [df.at[index, 'Periodo'], 'X']
                elif column == 'Uy':    T = [df.at[index, 'Periodo'], 'Y']
                
                # Calculo de k
                t = T[0]
                if t <= 0.5:        k = 1
                else:               k = min(0.75 + 0.5*t, 2)
                
                # Calculo de coeficiente C
                if t <= Tp:         Cs = 2.5
                elif Tp < t <= Tl:  Cs = 2.5*Tp/t
                else:               Cs = 2.5*(Tp*Tl)/(t**2)
                
                # Determinacion de la direccion del eje aplicado
                if T[1] == 'X':     print('Para SExx -----> Cx = ' + str(round(Cs,3)) + ', kx = ' + str(round(k,3)))
                elif T[1] == 'Y':   print('Para SEyy -----> Cy = ' + str(round(Cs,3)) + ', ky = ' + str(round(k,3)))
        
        
    def Masas_Participativas(self):
        # Filtrado de tabla
        ret = self.SapModel.Results.Setup.DeselectAllCasesAndCombosForOutput()
        ret = self.SapModel.Results.Setup.SetCaseSelectedForOutput("Modal")
        
        # Extracción de las masas participativas de los modos de vibración
        NumberResults = 0; LoadCase = []; Period = []; Ux = []; Uy = []; SumUx = []; SumUy = []; R_z = []; SumRz = []
        [NumberResults, LoadCase, _, _, Period, Ux, Uy, _, SumUx, SumUy, _, _, _, R_z, _, _, SumRz, ret] = self.SapModel.Results.ModalParticipatingMassRatios(NumberResults, LoadCase, [], [], Period, Ux, Uy, [], SumUx, SumUy, [], [], [], R_z, [], [], SumRz)
        
        # Creacion de tabla resumen
        Mode = np.arange(1, NumberResults + 1, 1)
        df = pd.DataFrame({'Case': LoadCase, 'Modo': Mode, 'Periodo': Period, 'Ux': Ux, 'Uy': Uy, 'Rz': R_z, 'SumUx': SumUx, 'SumUy': SumUy, 'SumRz': SumRz})
        data = df.round({'Periodo': 3, 'Ux': 3, 'Uy': 3, 'Rz': 3, 'SumUx': 3, 'SumUy': 3, 'SumRz':3})
        data.index = [''] * len(data)
        return data
        
    def Filter_Table(self, Table_name, Select_Display):
        # Seleccionar las cargas según Cases y Combinations
        if Select_Display[1] == 'Cases':
            ret = self.SapModel.DatabaseTables.SetLoadCasesSelectedForDisplay(Select_Display[0])
            ret = self.SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay([])
        elif Select_Display[1] == 'Combination':
            ret = self.SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay(Select_Display[0])
            ret = self.SapModel.DatabaseTables.SetLoadCasesSelectedForDisplay([])

        # Extraccion de resultados del modelo ETABS
        GroupName = Table_name
        FieldsKeysIncluded = []
        TableData = []
        [_, _, FieldsKeysIncluded, _, TableData, _] = self.SapModel.DatabaseTables.GetTableForDisplayArray(Table_name, [], GroupName, 0, FieldsKeysIncluded, 0, TableData)
        
        # Creacion de tabla filtrada
        n = len(FieldsKeysIncluded)
        output = [list(TableData[i:i+n]) for i in range(0, len(TableData), n)]
        m = len(output)
        Table = np.array(output).reshape(m,n)
        df = pd.DataFrame(Table, columns = FieldsKeysIncluded)
        return df
    
    def Irregularidad_Rigidez(self, OutputCase):
        # Filtrado de tabla
        df = self.Filter_Table('Story Stiffness', [['SExx', 'SEyy'], 'Cases'])
        df = df[df['OutputCase'] == OutputCase]
        if OutputCase =='SExx':
            CaseStiff = 'StiffX'
        else:
            CaseStiff = 'StiffY'
        
        # Conversión de string a float
        df[CaseStiff] = pd.to_numeric(df[CaseStiff])
               
        # Calculo de ratios para verificacion de irregularidad
        Ratio = pd.DataFrame(index=df.index, columns=['70% Ki+1', '80% Kprom', '60% Ki+1', '70% Kprom'])
        for i in range(1, len(df)):
            Ratio.at[df.index[i], '70% Ki+1'] = df.iloc[i][CaseStiff] / (0.7 * df.iloc[i-1][CaseStiff])
            Ratio.at[df.index[i], '60% Ki+1'] = df.iloc[i][CaseStiff] / (0.6 * df.iloc[i-1][CaseStiff])

        for i in range(3, len(df)):
            avg_3 = (df.iloc[i-1][CaseStiff] + df.iloc[i-2][CaseStiff] + df.iloc[i-3][CaseStiff]) / 3
            Ratio.at[df.index[i], '80% Kprom'] = df.iloc[i][CaseStiff] / (0.8 * avg_3)
            Ratio.at[df.index[i], '70% Kprom'] = df.iloc[i][CaseStiff] / (0.7 * avg_3)
        
        # Creacion de tabla resumen  
        data = pd.concat([df.reset_index(drop=True), Ratio.reset_index(drop=True)], axis=1)
        data.fillna(0, inplace=True)
        data = data.infer_objects(copy=False) # Quita comentario de advertencia
        data = data[data['Story'].notna()]
        data = data.round({CaseStiff: 2, '70% Ki+1': 2, '80% Kprom': 2, '60% Ki+1': 2, '70% Kprom': 2})
        data = data[['Story', 'OutputCase', CaseStiff, '70% Ki+1', '80% Kprom', '60% Ki+1', '70% Kprom']]
        data.index = [''] * len(data)
        return data
    
    def Irregularidad_Resistencia(self, OutputCase):
        # Filtrado de tabla
        df = self.Filter_Table('Story Stiffness', [['SExx', 'SEyy'], 'Cases'])
        df = df[df['OutputCase'] == OutputCase].copy()
        if OutputCase =='SExx':
            CaseShear = 'ShearX'
        else:
            CaseShear = 'ShearY'
        
        # Calculo de ratios para verificacion de irregularidad
        df[CaseShear] = pd.to_numeric(df[CaseShear])
        df['Resistencia (80%)'] = 0.0
        df['Ext. Resistencia (65%)'] = 0.0
        for i in range(len(df) - 1):
            df.at[df.index[i], 'Resistencia (80%)'] = df.at[df.index[i+1], CaseShear] / (0.8 * df.at[df.index[i], CaseShear])
            df.at[df.index[i], 'Ext. Resistencia (65%)'] = df.at[df.index[i+1], CaseShear] / (0.65 * df.at[df.index[i], CaseShear])

        # Creacion de tabla resumen
        data = df.round({CaseShear: 2, 'Resistencia (80%)': 2, 'Ext. Resistencia (65%)': 2})
        data = data[['Story', 'OutputCase', CaseShear, 'Resistencia (80%)', 'Ext. Resistencia (65%)']]
        data.index = [''] * len(data)
        return data
    
    def Irregularidad_Masa(self, Combination):
        # Filtrado de tabla 
        ret = self.SapModel.DatabaseTables.SetLoadCombinationsSelectedForDisplay([Combination])
        df = self.Filter_Table('Story Forces', [[Combination], 'Combination'])
        df = df[df['Location'] == 'Bottom']
        
        # Calculo de pesos por entrepiso
        df['P'] = pd.to_numeric(df['P'])
        df_copy = df.copy()
        df_copy['Peso por piso (kgf)'] = df_copy['P'].diff()
        df_copy.iloc[0, df_copy.columns.get_loc('Peso por piso (kgf)')] = df_copy.iloc[0, df_copy.columns.get_loc('P')]
        
        # Calculo de ratios para verificacion de irregularidad
        df_copy['Wi/Wi+1'] = float('nan')
        df_copy['Wi+1/Wi'] = float('nan')
        for i in range(len(df_copy) - 1):
            if df_copy.iloc[i + 1]['Peso por piso (kgf)'] != 0:
                df_copy.iloc[i, df_copy.columns.get_loc('Wi/Wi+1')] = df_copy.iloc[i + 1]['Peso por piso (kgf)'] / df_copy.iloc[i]['Peso por piso (kgf)']
            if df_copy.iloc[i]['Peso por piso (kgf)'] != 0:
                df_copy.iloc[i + 1, df_copy.columns.get_loc('Wi+1/Wi')] = df_copy.iloc[i]['Peso por piso (kgf)'] / df_copy.iloc[i + 1]['Peso por piso (kgf)']

        # Creacion de tabla resumen
        data = df_copy.round({'P': 2, 'Peso por piso (kgf)': 2, 'Wi/Wi+1': 2, 'Wi+1/Wi': 2})
        data = data.rename(columns={'P': 'Peso (kgf)'})
        data = data[['Story', 'OutputCase', 'Location', 'Peso (kgf)', 'Peso por piso (kgf)', 'Wi/Wi+1', 'Wi+1/Wi']]
        data.index = [''] * len(data)
        return data

    def Irregularidad_Torsional(self, OutputCase):
        # Filtrado de tabla 
        df = self.Filter_Table('Diaphragm Max Over Avg Drifts', [['RIX', 'RIY'], 'Combination'])
        if OutputCase == 'RIX':     eje = 'X'
        else:                       eje = 'Y'
        df = df[(df['StepType'] == 'Max') & (df['OutputCase'] == OutputCase) & (df['Item'].str.endswith(eje))]
        
        # Verificación de los ratios
        df['Ratio'] = pd.to_numeric(df['Ratio'])
        df['Ratio > 1.3'] = df['Ratio'].apply(lambda x: 'SI' if x > 1.3 else 'NO')
        df['Ratio > 1.5'] = df['Ratio'].apply(lambda x: 'SI' if x > 1.5 else 'NO')
        
        # Creacion de tabla resumen
        data = df[['Story', 'OutputCase', 'StepType', 'Item', 'Ratio', 'Ratio > 1.3', 'Ratio > 1.5']]
        data.index = [''] * len(data)
        return data
    
    def Derivas(self, OutputCase):
        # Filtrado de tabla
        df = self.Filter_Table('Story Drifts', [['RIX', 'RIY'],'Combination'])
        if OutputCase == 'RIX':     eje = 'X'
        else:                       eje = 'Y'
        df = df[(df['StepType'] == 'Max') & (df['OutputCase'] == OutputCase) & (df['Direction'] == eje)]
        
        # Verificación de los ratios
        df['Drift'] = pd.to_numeric(df['Drift'])
        df['Condicion'] = df['Drift'].apply(lambda x: 'CUMPLE' if x < 0.007 else 'NO CUMPLE')
        
        # Creacion de tabla resumen
        data = df[['Story', 'OutputCase', 'StepType', 'Drift', 'Z', 'Condicion']]
        data.index = [''] * len(data)        
        return data
    
    def Derivas_Grafico(self, OutputCase):       
        # Inputs para el ploteo
        data = self.Derivas(OutputCase)
        limit = [0.007] * (len(data) + 1)
        Direction = 'X - X' if OutputCase == 'RIX' else 'Y - Y'
        Z = ['BASE'] + ['NIVEL ' + str(i+1) for i in range(len(data))]
        Drift = [0] + [float(row['Drift']) for _, row in data.iterrows()]
        nivel_max = Drift.index(max(Drift))
        
        # Ploteo de gráfica de derivas
        plt.figure(figsize=(5,8))
        plt.plot(Drift, Z, 'b-o', label='Máx Drift = ' + str(round(max(Drift),3)))
        plt.plot(limit, Z, 'r--', label='Limite Drift E.030')
        plt.title('Derivas máximas ' + str(Direction))
        plt.grid(color='silver')
        plt.xlim(0,0.01)
        plt.ylim('BASE',Z[-1])
        legend = plt.legend(loc="lower right")
        legend.get_frame().set_alpha(1)
        plt.annotate(r'Max Drift', xy=(max(Drift), Z[nivel_max]), xycoords='data', xytext=(max(Drift)-1e-3, Z[nivel_max]), 
                     fontsize=10, arrowprops=dict(arrowstyle="->", connectionstyle="arc3,rad=.3"))
        return plt.show()
    
    def Sistema_Estructural(self, OutputCase):
        # Filtrado de tabla para columnas
        df1 = self.Filter_Table('Element Forces - Columns', [OutputCase,'Cases'])
        df1 = df1[(df1['Story'] == 'T_01') & (df1['Station'] == '0')]
        
        # Filtrado de tabla para muros
        df2 = self.Filter_Table('Pier Forces', [OutputCase, 'Cases'])
        if df2.empty:
            sum_wallx = 0
            sum_wally = 0
        else:
            df2 = df2[(df2['Story'] == 'T_01') & (df2['Location'] == 'Bottom')]
        
        # Calculo de cortante basal en columnas y muros
        for case in OutputCase:
            df1_copy = df1[df1['OutputCase'] == case].copy()
            df1_copy[['V2', 'V3']] = df1_copy[['V2', 'V3']].apply(pd.to_numeric, errors='coerce')
            if case == 'SExx':      sum_columnx = df1_copy['V2'].sum()
            elif case == 'SEyy':    sum_columny = df1_copy['V3'].sum()
            
            if not df2.empty:
                df2_copy = df2[df2['OutputCase'] == case].copy()
                df2_copy[['V2','V3']] = df2_copy[['V2','V3']].apply(pd.to_numeric, errors='coerce')
                if case == 'SExx':      sum_wallx = df2_copy['V2'].sum()
                elif case == 'SEyy':    sum_wally = df2_copy['V3'].sum()
                
        # Calculo de porcentajes de columnas y muros por eje
        total_x = abs(sum_columnx) + abs(sum_wallx)
        total_y = abs(sum_columny) + abs(sum_wally)
        percent_columnx = (abs(sum_columnx) / total_x) * 100
        percent_wallx = (abs(sum_wallx) / total_x) * 100
        percent_columny = (abs(sum_columny) / total_y) * 100
        percent_wally = (abs(sum_wally) / total_y) * 100
                
        # Ploteo de gráfica de porcentajes
        fig, ax = plt.subplots()
        r1 = np.arange(2)
        bar1_x = ax.bar(r1[0], percent_columnx, color='blue', edgecolor='grey', width=0.4, label='Columnas X-X')
        bar2_x = ax.bar(r1[0], percent_wallx, bottom=percent_columnx, color='lightblue', edgecolor='grey', width=0.4, label='Muros X-X')
        bar1_y = ax.bar(r1[1], percent_columny, color='green', edgecolor='grey', width=0.4, label='Columnas Y-Y')
        bar2_y = ax.bar(r1[1], percent_wally, bottom=percent_columny, color='lightgreen', edgecolor='grey', width=0.4, label='Muros Y-Y')
        for bar in bar1_x + bar2_x + bar1_y + bar2_y:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2.0, bar.get_y() + height / 2.0,
                    f'{height:.1f}%', ha='center', va='center', color='white', fontweight='bold')
        ax.set_xlabel('Ejes')
        ax.set_ylabel('Porcentaje (%)')
        ax.set_title('Cortante Basal en elementos estructurales (%)')
        ax.set_xticks(r1)
        ax.set_xticklabels(['Eje X-X', 'Eje Y-Y'])
        ax.legend()
        return plt.show()
    
    def Excel_tables(self):
        # Definicion de dataframes de cada análisis de irregularidad
        df1x = self.Irregularidad_Rigidez('SExx')
        df1y = self.Irregularidad_Rigidez('SEyy')
        df2x = self.Irregularidad_Resistencia('SExx')
        df2y = self.Irregularidad_Resistencia('SEyy')
        df3x = self.Irregularidad_Torsional('RIX')
        df3y = self.Irregularidad_Torsional('RIY')
        df4 = self.Irregularidad_Masa('P = CM + 25%CV')
        df5 = self.Masas_Participativas()
        df6x = self.Derivas('RIX')
        df6y = self.Derivas('RIY')
        
        # Creación de Excel de tablas de análisis sísmico
        dataframes = {
            'Irregularidad Rigidez': [df1x,df1y],
            'Irregularidad Resistencia': [df2x,df2y],
            'Irregularidad Torsional': [df3x,df3y],
            'Irregularidad Masa': df4,
            'Masas Participativas': df5,
            'Derivas': [df6x, df6y]
        }      
        book = Workbook()
        for sheet_name, df_pair in dataframes.items():
            sheet = book.create_sheet(title=sheet_name)
            start_row = 1
            
            if isinstance(df_pair, list):
                for df in df_pair:
                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                        for c_idx, value in enumerate(row, start=1):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                    start_row += df.shape[0] + 2
            
            else:
                for r_idx, row in enumerate(dataframe_to_rows(df_pair, index=False, header=True), start=start_row):
                    for c_idx, value in enumerate(row, start=1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

        # Eliminar la hoja de inicio por defecto
        if 'Sheet' in book.sheetnames:
            book.remove(book['Sheet'])
        book.save('Analisis_Sismico.xlsx') 
